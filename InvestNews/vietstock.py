#!/usr/bin/env python3
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.select import Select
import time
from datetime import datetime
from icecream import ic
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill


def enter_news_search(keyword):
    options = Options()
    # options.headless = True
    driver = webdriver.Chrome(options=options)
    # driver = webdriver.Chrome()
    link = 'https://vietstock.vn/'
    driver.get(link)

    # turn off popup advertisement
    wait = WebDriverWait(driver, 10)
    wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, "/html/body/div[3]/div[9]/div/div/div/a"))
    ).click()

    # click search box
    driver.find_element(
        By.XPATH, "/html/body/div[3]/div[4]/div/div[1]/div[3]/form/div/div/input").click()

    # enter string to search box
    wait.until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="popup-search-txt"]'))
    ).send_keys(keyword)

    # click to tag "Tin tuc"
    driver.find_element(
        By.XPATH, "/html/body/div[4]/div[2]/div/div[3]/ul/li[6]/span").click()

    # click to the first result
    time.sleep(2)
    driver.find_element(
        By.XPATH, "/html/body/div[4]/div[2]/div/div[4]/div/div[5]/div[2]/div/div/ul/li[1]/a/div[2]/div").click()

    # Export to mardown file
    export_text(driver)

    driver.quit()


def enter_news(self):
    driver = webdriver.Chrome()
    link = 'https://vietstock.vn/chung-khoan/co-phieu.htm'
    driver.get(link)

    # click the first link (be careful cuz the first link only true if you click on time < 8:00 AM)
    driver.find_element(
        By.XPATH, '//*[@id="channel-container"]/section[1]/div/div/div[2]/h2').click()

    # Export to mardown file
    export_text(driver)

    driver.quit()


def export_text(driver):
    now = datetime.now()
    date_now = now.strftime("%d-%m-%Y")
    file_path = '/media/dat/DISK/Dev/Automation/InvestNews/news/' + date_now + '.md'

    # body text
    content = driver.find_elements(By.ID, "vst_detail")

    # create and write file
    with open(file_path, 'w') as file:
        file.write(driver.current_url + '\n')
        for p in content:
            text = p.text.replace(' >>>', '\n>>>')
            file.write(text + '\n')

    # read file
    lines = []
    with open(file_path, 'r') as file:
        lines = file.readlines()

    # write file (remember for loop start at 0)
    with open(file_path, 'w') as file:
        for line_number, line_text in enumerate(lines):
            # remove line 2, line n, line n-1 (remember for loop start at 0)
            if line_number not in [1, len(lines)-1, len(lines)-2]:
                file.write(line_text)


def check_time():
    now = datetime.now()
    limit = datetime(now.year, now.month, now.day, 8, 0, 0)
    if now >= limit:
        print(f"It's more than {limit}")
        timkiem = [
            'Doc gi truoc gio giao dich'
        ]
        enter_news_search(timkiem[0])
        return
    enter_news()

# find stock downtrend in n phien
def filter(n_phien):
    options = Options()
    options.headless = True
    driver = webdriver.Chrome(options=options)
    tool_link = 'https://finance.vietstock.vn/bo-loc-co-phieu/in-decrease/0-0'
    driver.get(tool_link)

    # Select San giao dich
    select_element = driver.find_element(
        By.XPATH, '//*[@id="select-exchange"]')
    select_object = Select(select_element)
    select_object.select_by_visible_text('HOSE')

    # select giam(0) hay tang (1)
    select_element = driver.find_element(By.XPATH, '//*[@id="ddlOrderBy"]')
    select_object = Select(select_element)
    select_object.select_by_index(0)

    # Type in so_phien
    so_phien = driver.find_element(
        By.XPATH, '/html/body/div[2]/div[11]/div/div[2]/div/div/div/div/div[1]/div[1]/div/div[2]/div/input')
    so_phien.clear()
    so_phien.send_keys(str(n_phien))

    # Click xem_button
    driver.find_element(
        By.XPATH, '/html/body/div[2]/div[11]/div/div[2]/div/div/div/div/div[1]/div[2]/button').click()

    # wait for load search
    wait = WebDriverWait(driver, 10)
    wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="indecrease"]/table/tbody/tr/td[3]/a'))
    )

    export_excel(driver)

    driver.quit()


def export_excel(driver):
    now = datetime.now()
    date_now = now.strftime("%d-%m-%Y") 
    file_path = '/media/dat/DISK/Dev/Automation/InvestNews/news/' + date_now + '.xlsx'

    table_list = []

    # take table's header info
    table_header = driver.find_elements(
        By.XPATH, '//*[@id="indecrease"]/table/thead/tr/th')
    header_list = []
    for i in table_header:
        header_list.append(i.text)

    table_list.append(header_list)

    # take table's body info
    table_body = driver.find_elements(
        By.XPATH, '//*[@id="indecrease"]/table/tbody/tr')
    for item in table_body:
        a = item.text.replace(',', '.')  # change "4,5"-> "4.5"
        table_list.append(a.split())  # change string to list

    # read laiban.md file and create shit_stock list to filter the shit stock co tien su lai ban
    shit_stock = []
    laiBan_path = '/media/dat/DISK/Invest/laiBan.md'
    with open(laiBan_path, 'r') as file:
        shit_stock = file.read()

    shit_stock = shit_stock.split('\n')

    # color of cell
    fill_pattern = PatternFill(patternType='solid', fgColor='ff6054')

    # create new file .xlsv
    wb = Workbook()
    wb.save(file_path)

    wb = load_workbook(file_path)  # load file
    # ws = wb.active # active all sheet
    ws = wb['Sheet']  # active 1 sheet

    table_row = len(table_list)
    table_col = len(table_list[0])

    for row in range(1, table_row+1):
        for col in range(1, table_col+1):
            char = get_column_letter(col)
            ws[char+str(row)].value = table_list[row-1][col-1]
            # fill color if that stock is in shit stock
            if char == 'C' and table_list[row-1][col-1] in shit_stock:
                ws[char+str(row)].fill = fill_pattern

    # resize column width
    column_widths = []
    for row in table_list:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(cell) > column_widths[i]:
                    column_widths[i] = len(cell)
            else:
                column_widths += [len(cell)]

    for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
        ws.column_dimensions[get_column_letter(i)].width = column_width

    # set zoom in
    ws.sheet_view.zoomScale = 150

    wb.save(file_path)
