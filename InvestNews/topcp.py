#!/usr/bin/env python3
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.select import Select

import time
from datetime import datetime
from icecream import ic
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

def filter():
  options = Options()
  options.headless = True
  driver = webdriver.Chrome(options=options)
  vietstock_tool = 'https://finance.vietstock.vn/bo-loc-co-phieu/in-decrease/0-0'
  driver.get(vietstock_tool)

  # Select San giao dich
  select_element = driver.find_element(By.XPATH,'//*[@id="select-exchange"]')
  select_object = Select(select_element)
  select_object.select_by_visible_text('HOSE')

  # select giam(0) hay tang (1)
  select_element = driver.find_element(By.XPATH,'//*[@id="ddlOrderBy"]')
  select_object = Select(select_element)
  select_object.select_by_index(0)

  # Type in so_phien
  so_phien = driver.find_element(By.XPATH, '/html/body/div[2]/div[11]/div/div[2]/div/div/div/div/div[1]/div[1]/div/div[2]/div/input')
  so_phien.clear()
  so_phien.send_keys("5")

  # Click xem_button
  xem_button = driver.find_element(By.XPATH, '/html/body/div[2]/div[11]/div/div[2]/div/div/div/div/div[1]/div[2]/button')
  xem_button.click()

  # wait for load search
  wait = WebDriverWait(driver, 10)
  wait.until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="indecrease"]/table/tbody/tr/td[3]/a'))
  )

  export_excel(driver)

  driver.quit()

def export_excel(driver):
  now = datetime.now()
  date_now = now.strftime("%d-%m-%Y") + '.xlsx'
  file_path = './news/' + date_now

  table_list = []

  # take table's header info
  table_header = driver.find_elements(By.XPATH, '//*[@id="indecrease"]/table/thead/tr/th')
  header_list = []
  for i in table_header:
    header_list.append(i.text)

  table_list.append(header_list)
  
  # take table's body info
  table_body = driver.find_elements(By.XPATH, '//*[@id="indecrease"]/table/tbody/tr')
  for item in table_body:
    a = item.text.replace(',', '.') # change "4,5"-> "4.5"
    table_list.append(a.split()) # change string to list 

  # read laiban.md file and create shit_stock list to filter the shit stock co tien su lai ban
  shit_stock = []
  with open('./laiBan.md', 'r') as file:
    shit_stock = file.read()

  shit_stock = shit_stock.split('\n')

  # color of cell
  fill_pattern = PatternFill(patternType='solid', fgColor='ff6054')

  # create new file .xlsv
  wb = Workbook()
  wb.save(file_path)

  wb = load_workbook(file_path) # load file
  # ws = wb.active # active all sheet
  ws = wb['Sheet'] # active 1 sheet

  table_row = len(table_list)
  table_col = len(table_list[0])

  for row in range(1, table_row+1):
    for col in range(1, table_col+1):
      char = get_column_letter(col)
      ws[char+str(row)].value = table_list[row-1][col-1]
      # fill  color if that stock is shit stock
      if char == 'C' and table_list[row-1][col-1] in shit_stock:
        ws[char+str(row)].fill = fill_pattern

  # resize column width
  column_widths = []
  for row in table_list:
    for i, cell in enumerate(row):
      if len(column_widths) > i:
        if len(cell) > column_widths[i]:
          column_widths[i] = len(cell)
      else:
        column_widths += [len(cell)]
      
  for i, column_width in enumerate(column_widths,1):  # ,1 to start at 1
    ws.column_dimensions[get_column_letter(i)].width = column_width

  # set zoom in
  ws.sheet_view.zoomScale = 150

  wb.save(file_path)
  
if __name__ == '__main__':
  filter()

# schedule.every().monday.tuesday.wednesday.thursday.friday.at("08:01").do(vietstock, keyword="Top co phieu dang chu y")
